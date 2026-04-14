#!/usr/bin/env python3
"""
build.py — Reads Tabla_Positioning.xlsx and generates data.json
Run this every time you update the Excel, then push to GitHub.

Usage:
    python build.py
"""

import json
import openpyxl
import sys
import os

EXCEL_FILE = "Tabla Positioning.xlsx"
OUTPUT_FILE = "data.json"

# Map hex colors to recommendation levels
COLOR_MAP = {
    "FF00B050": "bullish",    # Green
    "00B050":   "bullish",
    "FFFFC000": "neutral",    # Yellow/Amber
    "FFC000":   "neutral",
    "FFC00000": "bearish",    # Red
    "C00000":   "bearish",
}

# Change arrow mapping
CHANGE_MAP = {
    "↑": "up",
    "↓": "down",
    "-": "unchanged",
    None: "unchanged",
    "": "unchanged",
}

def get_recommendation_color(cell):
    """Extract the recommendation level from the dot's font color."""
    if cell.value != "●":
        return None
    color = cell.font.color
    if color and color.rgb:
        rgb = str(color.rgb).replace("FF", "", 1) if len(str(color.rgb)) == 8 else str(color.rgb)
        full_rgb = str(color.rgb)
        # Try both with and without leading FF
        for key in [full_rgb, rgb]:
            if key in COLOR_MAP:
                return COLOR_MAP[key]
    return "neutral"  # fallback

def build_data():
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: {EXCEL_FILE} not found in current directory.")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # Try Spanish sheet first, fall back to first sheet
    if "20260413 (spanish)" in wb.sheetnames:
        ws = wb["20260413 (spanish)"]
    else:
        ws = wb[wb.sheetnames[-1]]  # Use last sheet (usually Spanish)

    # Find the sheet name to extract date
    sheet_name = ws.title
    date_part = sheet_name.split(" ")[0] if " " in sheet_name else sheet_name

    # Parse the structured data
    # The Spanish sheet layout (columns B-I, 1-indexed):
    # Row 9:  Headers (Recom. | Cambio | Principales Factores Macro | Instrumento)
    # Row 10: CER (category) with dot
    # Row 11-13: Corto/Medio/Largo for CER
    # Row 14: empty
    # Row 15: Tasa Fija (category) with dot
    # Row 16-18: Corto/Medio/Largo for Tasa Fija
    # Row 19: empty
    # Row 20: Tipo de Cambio (category) with dot
    # Row 21: empty
    # Row 22: Curva en Dólares (category) with dot
    # Row 23-25: Corto/Medio/Largo for Curva en Dólares

    categories = []

    # Helper: read a category block
    def read_category(cat_row, sub_rows, cat_name):
        cat_cell_recom = ws.cell(row=cat_row, column=6)  # F = recommendation dot
        cat_cell_change = ws.cell(row=cat_row, column=7)  # G = change
        cat_cell_macro = ws.cell(row=cat_row, column=8)   # H = macro factors
        cat_cell_instr = ws.cell(row=cat_row, column=9)   # I = instrument

        category = {
            "name": cat_name,
            "recommendation": get_recommendation_color(cat_cell_recom),
            "change": CHANGE_MAP.get(cat_cell_change.value, "unchanged"),
            "macro": cat_cell_macro.value or "",
            "instrument": cat_cell_instr.value or "",
            "subcategories": []
        }

        for sr in sub_rows:
            sub_name_cell = ws.cell(row=sr, column=4)  # D = sub name
            sub_recom_cell = ws.cell(row=sr, column=6)
            sub_change_cell = ws.cell(row=sr, column=7)
            sub_macro_cell = ws.cell(row=sr, column=8)
            sub_instr_cell = ws.cell(row=sr, column=9)

            if sub_name_cell.value:
                sub = {
                    "name": sub_name_cell.value,
                    "recommendation": get_recommendation_color(sub_recom_cell),
                    "change": CHANGE_MAP.get(sub_change_cell.value, "unchanged"),
                    "macro": sub_macro_cell.value or "",
                    "instrument": sub_instr_cell.value or "",
                }
                category["subcategories"].append(sub)

        return category

    categories.append(read_category(10, [11, 12, 13], "CER"))
    categories.append(read_category(15, [16, 17, 18], "Tasa Fija"))
    categories.append(read_category(20, [], "Tipo de Cambio"))
    categories.append(read_category(22, [23, 24, 25], "Curva en Dólares"))

    data = {
        "date": date_part,
        "categories": categories,
    }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"✅ Generated {OUTPUT_FILE} from {EXCEL_FILE}")
    print(f"   Date: {date_part}")
    print(f"   Categories: {len(categories)}")
    for cat in categories:
        print(f"     - {cat['name']} ({cat['recommendation']}) [{len(cat['subcategories'])} sub]")

if __name__ == "__main__":
    build_data()
